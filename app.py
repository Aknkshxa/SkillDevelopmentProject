from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from PIL import Image
import pytesseract
import tabula

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf', 'jpeg', 'jpg', 'png', 'xlsx'}

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_pdf(file_path):
    dfs = tabula.read_pdf(file_path, pages="all", multiple_tables=True)
    if not dfs:
        return None
    return pd.concat(dfs, ignore_index=True)

def process_image(file_path):
    try:
        img = Image.open(file_path)
        text = pytesseract.image_to_string(img)
        data = [line.split() for line in text.splitlines() if line.strip()]
        return pd.DataFrame(data)
    except Exception as e:
        print(f"Error processing image: {e}")
        return None

def process_excel(file_path):
    try:
        excel_data = pd.read_excel(file_path, sheet_name=None)
        return pd.concat(excel_data.values(), ignore_index=True)
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        return None

def read_input_file(file_path):
    file_extension = os.path.splitext(file_path)[1].lower()
    if file_extension == '.pdf':
        return process_pdf(file_path)
    elif file_extension in ['.jpeg', '.jpg', '.png']:
        return process_image(file_path)
    elif file_extension == '.xlsx':
        return process_excel(file_path)
    else:
        return None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process_file():
    if 'file' not in request.files or 'subjectName' not in request.form:
        return jsonify({'message': 'File or subject name missing'}), 400

    print(request.form)
    print(request.files)

    file = request.files['file']
    subject_name = request.form.get('subjectName')

    if not file or not allowed_file(file.filename):
        return jsonify({'message': 'Invalid file type'}), 400

    filename = file.filename
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(file_path)

    combined_df = read_input_file(file_path)
    if combined_df is None:
        return jsonify({'message': 'No valid data found in the file'}), 400

    otp = "output_combined.xlsx"
    with pd.ExcelWriter(otp, engine='openpyxl') as writer:
        combined_df.to_excel(writer, sheet_name="Combined_Data", index=False)

    wb = load_workbook(otp)
    ws = wb['Combined_Data']
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    subject_column = None
    student_name_column = 1
    header_row = 4
    subject_name = subject_name.lower()
    for cell in ws[header_row]:
        if cell.value and isinstance(cell.value, str):
            header_value = cell.value.strip().lower()
            if subject_name == header_value:
                subject_column = cell.column
                break
    if subject_column is None:
        return jsonify({'message': f'Subject {subject_name} not found'}), 400

    highlighted = False
    for row in ws.iter_rows(min_row=header_row + 1, min_col=subject_column, max_col=subject_column):
        attendance_cell = row[0]
        try:
            if isinstance(attendance_cell.value, str):
                attendance_value = float(attendance_cell.value.replace('%', '').replace(',', '').strip())
            else:
                attendance_value = float(attendance_cell.value)

            if attendance_value < 80:
                student_name_cell = ws.cell(row=attendance_cell.row, column=student_name_column)
                student_name_cell.fill = highlight_fill
                attendance_cell.fill = highlight_fill
                highlighted = True
        except ValueError:
            pass

    # Bar Chart Creation
    # Assuming the attendance values are in the column `subject_column` and starting from `header_row+1`
    chart = BarChart()
    chart.title = f"Attendance Bar Chart - {subject_name.capitalize()}"
    chart.x_axis.title = "Students"
    chart.y_axis.title = "Attendance (%)"

    data = Reference(ws, min_col=subject_column, min_row=header_row, max_row=ws.max_row)
    categories = Reference(ws, min_col=student_name_column, min_row=header_row + 1, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart_location = f"{get_column_letter(subject_column + 3)}{header_row + 5}"
    ws.add_chart(chart, chart_location)

    output_file = "output_highlighted_with_chart.xlsx"
    wb.save(output_file)

    return jsonify({'message': f'Highlighting and chart creation completed. Check {output_file}'}), 200

if __name__ == '__main__':
    app.run(debug=True)
